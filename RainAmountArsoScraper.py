from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import calendar
import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



url = "https://meteo.arso.gov.si/met/sl/app/webmet/#webmet==8Sdwx2bhR2cv0WZ0V2bvEGcw9ydlJWblR3LwVnaz9SYtVmYh9iclFGbt9SaulGdugXbsx3cs9mdl5WahxXYyNGapZXZ8tHZv1WYp5mOnMHbvZXZulWYnwCchJXYtVGdlJnOn0UQQdSf;"

chromeOptions = Options()
driver = webdriver.Chrome()
year = 2021


def is_leap_year(year):
    return calendar.isleap(year)


def get_no_of_days(year):
    if is_leap_year(year):
        return 366
    return 365

def init_stuff():
    driver.get(url)
    dt = WebDriverWait(driver, 10).until(EC.visibility_of_any_elements_located((By.CLASS_NAME, "academa-archive-fieldset-content")))
    data_type = dt[0]
    driver.implicitly_wait(5)
    daily_data = data_type.find_elements(By.TAG_NAME, "table")[4]
    button = daily_data.find_element(By.CLASS_NAME, "academa-checkbox-button")
    button.click()
    station_type_container = driver.find_elements(By.CLASS_NAME, "academa-archive-fieldset-content")[1]

    station_types = station_type_container.find_elements(By.TAG_NAME, "table")

    for i in range(2, len(station_types), 2):
        station_types[i].find_element(By.CLASS_NAME, "academa-button").click()

    time_period = driver.find_elements(By.CLASS_NAME, "academa-archive-fieldset-content")[2]
    period_data = time_period.find_elements(By.TAG_NAME, "table")[3]
    button = period_data.find_element(By.CLASS_NAME, "academa-checkbox-button")
    button.click()

def write_col(ws, col, data_list):
    for index, item in enumerate(data_list, start=1):
        ws.cell(row=index, column=col, value=item)


def save_stuff(data):
    wb = Workbook()
    ws = wb.active

    dates = data[0][1]
    dates.insert(0, "Dates")
    write_col(ws, 1, dates)

    for s_no, (name, date, rain, snow) in enumerate(data, start=1):
        rain.insert(0, name)
        snow.insert(0, " ")
        write_col(ws, s_no*2, rain)
        write_col(ws, s_no*2+1, snow)

    wb.save("./data/raiaaaaan_"+str(year)+".xlsx")


def get_data_for_year():

    driver.implicitly_wait(2)
    spremenljivke = driver.find_elements(By.CLASS_NAME, "academa-archive-form-group")[0]
    s_body = spremenljivke.find_element(By.TAG_NAME, "tbody")
    ss = s_body.find_elements(By.XPATH, "./child::tr")
    all_daily = ss[4]
    all_daily_td = all_daily.find_elements(By.TAG_NAME, "td")
    td_button_vse_dnevne_met_spr = all_daily_td[0]
    td_button_vse_dnevne_met_spr.find_element(By.CLASS_NAME, "academa-checkbox-button").click()

    all_daily_options = ss[5]

    all_options_buttons_table = all_daily_options.find_element(By.TAG_NAME, "table")
    all_options_buttons_body = all_options_buttons_table.find_element(By.TAG_NAME, "tbody")

    all_options_buttons_trs = all_options_buttons_body.find_elements(By.XPATH, "./child::tr")

    # 24-urna količina padavin in višina snežne odeje
    all_options_buttons_trs[8].find_element(By.CLASS_NAME, "academa-button").click()
    all_options_buttons_trs[9].find_element(By.CLASS_NAME, "academa-button").click()

    driver.find_element(By.XPATH, "//td[text()='Postaje']").click()

    s1 = driver.find_element(By.ID, "academaMapData")
    s2 = s1.find_elements(By.XPATH, "./child::div")[1]
    s3 = s2.find_element(By.TAG_NAME, "table")
    s4 = s3.find_element(By.TAG_NAME, "tbody")
    s5 = s4.find_elements(By.XPATH, "./child::tr")[1]
    s6 = s5.find_element(By.TAG_NAME, "div")
    s7 = s6.find_elements(By.XPATH, "./child::table")[1]
    s8 = s7.find_element(By.TAG_NAME, "tbody")
    s9 = s8.find_elements(By.XPATH, "./child::tr")[5]
    s10 = s9.find_element(By.TAG_NAME, "fieldset")
    stations = s10.find_element(By.TAG_NAME, "tbody")

    tds = stations.find_elements(By.TAG_NAME, "td")

    scraped_data = []

    for station in tds:
        driver.execute_script("arguments[0].click();", station)
        dt = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "aarchive-table")))
        data_table = dt[0]

        data_table_rows = data_table.find_elements(By.TAG_NAME, "tr")

        name_and_coords = data_table_rows[0].find_elements(By.TAG_NAME, "td")[0].text.replace("\n", " ")

        rain = []
        snow = []
        date = []

        for i in range(1, len(data_table_rows)-1):

            row_data = data_table_rows[i].find_elements(By.TAG_NAME, "td")

            date.append(row_data[0].text)
            rain.append(row_data[1].text)
            snow.append(row_data[2].text)

        scraped_data.append((name_and_coords, date, rain, snow))
        driver.find_element(By.XPATH, "//td[text()='Postaje']").click()


    save_stuff(scraped_data)


init_stuff()
start_date = datetime.datetime(year=year, month=1, day=1)
end_date = datetime.datetime(year=year, month=12, day=31)

# 0 because you need 2 digits for one digit days
start_date_string = str(start_date.year) + "-0" + str(start_date.month) + "-0" + str(start_date.day)
end_date_string = str(end_date.year) + "-" + str(end_date.month) + "-" + str(end_date.day)

input_start_date = driver.find_elements(By.TAG_NAME, "input")[7]
input_start_date.send_keys(start_date_string)

input_end_date = driver.find_elements(By.TAG_NAME, "input")[8]
input_end_date.send_keys(end_date_string)

button = driver.find_elements(By.CLASS_NAME, "academa-button1")[1]
button.click()

get_data_for_year()




