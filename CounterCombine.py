# vamos
import json
import sys
from datetime import datetime as dt
import datetime
import openpyxl
import pandas as pd
import xlsxwriter

import calendar
import numpy as np

data_paths = {
    "2021": {"hourly": "./data/traffic_2021.tab",
             "minutely": "./data/traffic_2021_5min.tab"
             },
    "2020": {"hourly": "./data/traffic_2020.tab",
             "minutely": "./data/traffic_2020_5min.tab"
             },
    "2019": {"hourly": "./data/traffic_2019.tab",
             "minutely": "./data/traffic_2019_5min.tab"
             }
}

data_paths_combined = {
    "2021": {
        "hourly": "./data/traffic_2021_combined.xlsx",
        "minutely": "./data/traffic_2021_combined_minutely.xlsx"
    },
    "2020": {"hourly": "./data/traffic_2020_combined.xlsx",
             "minutely": "./data/traffic_2020_combined_minutely.xlsx"
             },
    "2019": {
        "hourly": "./data/traffic_2019_combined.xlsx",
        "minutely": "./data/traffic_2019_combined_minutely.xlsx"
    }
}


class FileWriter:

    def save_stuff(self):
        self.workbook.close()


    # add -1s where values are missing (can happen for example, if there is no values for some counter in the last day,
    # therefore write_combined_values does not get called and no -1s are written at the end of this counter's column)
    def wrap_it_up(self, is_minutely):

        print("wrapping it up")

        for counter_name in self.counter_cols:
            row = self.counter_start_row[counter_name]
            counter_col = self.counter_cols[counter_name]

            while row < get_no_days_in_year(self.year)*get_values_for_day(is_minutely)+1:
                self.worksheet.write(row, counter_col, -1)
                row += 1
        print("done")


    def __init__(self, year, is_minutely):

        self.year = year
        self.is_minutely = is_minutely

        self.workbook = xlsxwriter.Workbook("./data/temp_worksheet.xlsx")
        self.worksheet = self.workbook.add_worksheet()

        self.worksheet = add_time_labels(self.worksheet, year, is_minutely)

        self.counter_cols = {}
        self.counter_start_row = {}

    def get_day(self, date):
        return date.split(".")[0]

    def get_month(self, date):
        return date.split(".")[1]

    def get_start_row(self, day, month):
        date_string = str(day)+"." + str(month)+"." + str(self.year)
        date_object = dt.strptime(date_string, '%d.%m.%Y')
        return (date_object.timetuple().tm_yday-1) * get_values_for_day(self.is_minutely) + 1

    def write_combined_values_to_file_hourly(self, values, counter_name, day, month):

        counter_name = counter_name[:-1]

        if counter_name not in self.counter_cols:
            self.counter_cols[counter_name] = self.worksheet.dim_colmax + 1
            self.worksheet.write(0, self.counter_cols[counter_name], counter_name)

        counter_col = self.counter_cols[counter_name]

        if counter_name not in self.counter_start_row:
            self.counter_start_row[counter_name] = 1

        start_row = self.get_start_row(day, month)
        row = self.counter_start_row[counter_name]
        while row < start_row:
            self.worksheet.write(row, counter_col, -1)
            row += 1
        self.counter_start_row[counter_name] = start_row

        # Write the values to the target columns
        for row, value in enumerate(values, start=start_row):
            self.worksheet.write(row, counter_col, value)
        self.counter_start_row[counter_name] = self.counter_start_row[counter_name] + len(values)

    # def write_combined_values_to_file_minutely(self, no_of_vehicles, counter_value, day, month, year):
    #
    #     try:
    #         time = dt.strptime("00:00:00", "%H:%M:%S")
    #         for row in range(0, len(no_of_vehicles)):
    #             self.file.write(get_row_string(no_of_vehicles, get_counter_value_no_lanes_from_value(counter_value),
    #                                            time.strftime("%H:%M:%S"), day, month, year, row))
    #             time += datetime.timedelta(minutes=5)
    #     except Exception as e:
    #         raise e

    def write_combined_values_to_file(self, no_of_vehicles, counter_value, day, month, year, is_minutely):
        if is_minutely:
            # todo
            pass
            # self.write_combined_values_to_file_minutely(no_of_vehicles, counter_value, day, month, year)
        else:
            self.write_combined_values_to_file_hourly(no_of_vehicles, counter_value, day, month)


def add_counter_names(worksheet, counter_cols):

    start_col = 2
    for ctr_name in counter_cols:
        worksheet.write(0, start_col, ctr_name)
        start_col += 1

    return worksheet


def replace_null_values(counter_cols, year, is_minutely):

    print("replace null values start")

    data = pd.read_excel("./data/temp_worksheet.xlsx", header=0)
    coords = get_coords_of_ctrs()

    print("init worksheet")
    workbook = xlsxwriter.Workbook(get_combined_path_from_year(year, is_minutely))
    worksheet = workbook.add_worksheet() # todo you have two worksheet: one for writing to temp and one for writing to final file with the sam
    worksheet = add_time_labels(worksheet, year, is_minutely)
    worksheet = add_counter_names(worksheet, counter_cols)
    print("end of init")

    null_values = {}
    i = 0

    for counter_name in counter_cols:

        print("replacing null values for counter: " + counter_name + ", index: " + str(i))
        i += 1

        null_values[counter_name] = []

        for row in range(0, get_no_days_in_year(year)*get_values_for_day(is_minutely)):
            value = data.iloc[row, counter_cols[counter_name]]

            if row == 2352 and year == 2021 and not is_minutely:
                # todo: better implementation
                # todo: do it for other years also
                # break cause there is no more data after 8th of April
                break

            if row < 5042 and year == 2019 and not is_minutely:
                null_values[counter_name].append((row, True))
                continue

            if value == -1:
                null_values[counter_name].append((row, True))
                worksheet.write(row+1, counter_cols[counter_name],
                                handle_null_value(data, row, counter_cols, counter_name,
                                    null_values, coords, is_minutely))
            elif value is not None:
                worksheet.write(row+1, counter_cols[counter_name], value)
                null_values[counter_name].append((row, False))
            else:
                break

    workbook.close()


def add_time_labels(worksheet, year, is_minutely):

    print("start init worksheet")
    start_date = dt(year=year, month=1, day=1)

    values_for_day = get_values_for_day(is_minutely)

    row_start = 1

    for i in range(get_no_days_in_year(year)):
        current_date = start_date + datetime.timedelta(days=i)
        worksheet.write(row_start, 0, current_date.strftime("%-d.%-m."))
        row_start += values_for_day

    row_start = 1
    time = dt(day=1, month=1, year=year, hour=0, minute=0)
    for i in range(get_no_days_in_year(year) * values_for_day):
        worksheet.write(row_start, 1, str(time.hour) + ":" + str(time.minute))
        row_start += 1
        time = increase_date_by(time, 1, is_minutely)

    print("init worksheet completed")

    return worksheet


def get_coords_of_ctrs():
    coords_file = json.load(open('./data/counters.json', 'r'))

    coords = {}

    for ctr_name, ctr_data in coords_file.items():

        lat, lon = get_lat_lan(ctr_data)

        # overrides counters on the same lane, doesn't matter since counters on the same lane
        # (btw this also holds for counters in the other direction) have the same coordinates
        coords[ctr_name[:-1]] = {'latitude': 0,
                                 'longitude': 0}
        coords[ctr_name[:-1]]['latitude'] = lat
        coords[ctr_name[:-1]]['longitude'] = lon

    return coords


def increase_date_by(date, amount, is_minutely):
    if is_minutely:
        return date + datetime.timedelta(minutes=amount*5)
    else:
        return date + datetime.timedelta(hours=amount)

def get_no_of_values_in_a_week(is_minutely):
    if is_minutely:
        return 24*12*7
    else:
        return 24*7


def get_values_from_prev_weeks(data, row, col, null_values_ctr, is_minutely):
    values_in_a_week = get_no_of_values_in_a_week(is_minutely)
    r = row - values_in_a_week
    limit = r - 5 * values_in_a_week
    if limit < 0:
        limit = 0
    while r > limit:
        value = data.iloc[r, col]
        if not null_values_ctr[r][1] and value != -1:
            return value
        r -= values_in_a_week
    return None


def find_coords_of_counter(coords, counter_name):
    for ctr_name, ctr_data in coords.items():
        if ctr_name[:-1] == counter_name:
            return ctr_data['latitude'], ctr_data['longitude']
    raise Exception("could not find coords of counter: " + counter_name)


def get_dist(ctr_data, lat, lon):
    lat_ctr = ctr_data['latitude']
    lon_ctr = ctr_data['longitude']
    return abs(lat_ctr-lat) + abs(lon_ctr-lon)


def get_lat_lan(ctr_data):
    lat = ctr_data['latitude']
    lon = ctr_data['longitude']
    return lat, lon


def get_value_from_counter(data, row, counter_name):
    return data.loc[:, counter_name].values.tolist()[row]


def get_value_from_closest_counter(data, coords, row, counter_cols, null_values, counter_name):

    if counter_name not in coords:
        return -1

    lat = coords[counter_name]['latitude']
    lon = coords[counter_name]['longitude']

    min_dist = sys.maxsize
    return_ctr_value = None

    values_for_day = data.iloc[row]

    for ctr_name, ctr_data in coords.items():
        dist = get_dist(ctr_data, lat, lon)

        value = -1
        try:
            value = values_for_day[counter_cols[ctr_name]]
        except:
            # no data for counter with name ctr_name
            continue

        if ctr_name != counter_name and min_dist > dist and value != -1 \
                and ctr_name in null_values and not null_values[ctr_name][row][1]:
            min_dist = dist
            return_ctr_value = value

    if return_ctr_value is not None:
        return return_ctr_value
    else:
        return -1
        # raise Exception("could not find closest counter with non null value")


def is_next_value_null(data, r, c):
    if data.iloc[r+1, c] == -1:
        return True
    return False


def handle_null_value(data, row, counter_cols, counter_name, null_values, coords, is_minutely):

    if row > 0 and not null_values[counter_name][row-1][1] and not is_next_value_null(data, row,
                                                                                      counter_cols[counter_name]):
        return data.iloc[row-1, counter_cols[counter_name]]
    else:
        value_from_prev_weeks = get_values_from_prev_weeks(data, row, counter_cols[counter_name],
                                                           null_values[counter_name], is_minutely)
        if value_from_prev_weeks is not None:
            return value_from_prev_weeks
        else:
            return get_value_from_closest_counter(data, coords, row, counter_cols, null_values, counter_name)


def get_values_for_day(is_minutely):
    if is_minutely:
        return 288
    return 24


def get_no_days_in_year(year):
    if calendar.isleap(year):
        return 366
    return 365


def get_counter_value_no_lanes_from_value(counter_value):
    return counter_value[:-1]


def get_combined_path_from_year(year, is_minutely):
    if not test_file:
        if not is_minutely:
            return data_paths_combined[str(year)]["hourly"]
        else:
            # return data_paths_combined[str(year)]["minutely"]
            pass
    else:
        # return "./data/final_test.tab"
        pass


def get_row_string(no_of_vehicles, counter_value, time, day, month, year, row):
    return str(year) + "\t" + str(month) + "\t" \
        + str(day) + "\t" + counter_value + "\t" + time \
        + "\t" + str(no_of_vehicles[row]) + "\n"


def get_vehicles_from_row(data, row):
    v = data.iloc[row, 5]
    if pd.isna(v):
        return -1
    else:
        return data.iloc[row, 5]


def get_hour_from_row(data, row):
    try:
        return data.iloc[row, 4]
    except:
        return -1


def get_daily_vehicles_by_day_for_counter(data, counter_ix):
    # print("get_daily_vehicles_by_hour_for_counter -> ")

    vehicles = []
    counter_value = get_counter_value(data, counter_ix)
    cix = counter_ix

    next_counter_value = get_counter_value(data, counter_ix)

    hour = 0
    number_of_values = 0

    while counter_value == next_counter_value:
        if get_hour_from_row(data, cix) == hour:
            vehicles.append(get_vehicles_from_row(data, cix))
            cix += 1
            number_of_values += 1
        else:
            vehicles.append(-1)
        hour += 1

        try:
            next_counter_value = get_counter_value(data, cix)
        except:
            break

    while len(vehicles) < 24:
        vehicles.append(-1)

    return vehicles, number_of_values


# ignore seconds
def is_equal_by_hour_and_minute(time1, time2):
    return time1[:5] == time2[:5]


def time_dif_between_in_mins(t1, t2):
    time1 = dt.strptime(t1, "%H:%M:%S")
    time2 = dt.strptime(t2, "%H:%M:%S")

    return (time1 - time2).total_seconds() / 60


def get_daily_vehicles_by_hour_for_counter(data, counter_ix):
    # print("get_daily_vehicles_by_hour_for_counter -> ")
    vehicles = []
    counter_value = get_counter_value(data, counter_ix)
    cix = counter_ix

    next_counter_value = get_counter_value(data, counter_ix)

    time = dt.strptime("00:00:00", "%H:%M:%S")
    number_of_values = 0

    # print("get_hour_from_row: " + counter_value + ", " + get_hour_from_row(data, cix) + ", day: " +
    # str(get_day_from_row(data, cix)))

    previous_time = None

    while counter_value == next_counter_value:
        # if counter_value == "0873-12":
        # print("hour_from_row: " + str(get_hour_from_row(data, cix)) + ", time:  " + time.strftime("%H:%M:%S"))

        if previous_time is not None and time_dif_between_in_mins(get_hour_from_row(data, cix), previous_time) < 5:
            cix += 1
            number_of_values += 1
            next_counter_value = get_counter_value(data, cix)
            continue
        else:
            if is_equal_by_hour_and_minute(get_hour_from_row(data, cix), time.strftime("%H:%M:%S")):
                vehicles.append(get_vehicles_from_row(data, cix))
                previous_time = get_hour_from_row(data, cix)
                cix += 1
                number_of_values += 1
            else:
                vehicles.append(0)

        time += datetime.timedelta(minutes=5)

        try:
            next_counter_value = get_counter_value(data, cix)
            # print("cix: " + str(cix) + ", " + next_counter_value + ", " + get_hour_from_row(data, cix))
        except:
            break

    while len(vehicles) < 288:  # = 12*24: 12 5-min intervals in one hour * 24 hours
        vehicles.append(0)

    return vehicles, number_of_values


def get_counter_value(data, counter_ix):
    # print("get_counter_value -> ")
    return data.iloc[counter_ix, 3]


def get_counter_value_test(data, counter_ix):
    # print("get_counter_value -> ")
    return data.iloc[counter_ix]


# returns name + direction
def get_counter_name_from_value(counter_value):
    try:
        return counter_value[:6]
    except:
        return ""


def get_counter_lane_from_value(counter_value):
    try:
        return counter_value[6:7]
    except:
        return ""


def combine_vehicle_numbers(v1, v2):
    # print("combine_vehicle_numbers -> ")

    for i in range(0, len(v1)):
        if v1[i] == -1:
            v1[i] = v2[i]
        elif v2[i] == -1:
            pass
        else:
            v1[i] += v2[i]
    return v1


def get_daily_vehicles_for_counter(data, counter_ix, is_minutely):
    if is_minutely:
        # return get_daily_vehicles_by_hour_for_counter(data, counter_ix)
        # todo doesnt work yet
        pass
    else:
        return get_daily_vehicles_by_day_for_counter(data, counter_ix)


# combines the counters that are on the same location and direction but on
# different lanes
def combine_counter(data, first_counter_row_num, day, month, year, is_minutely, file_writer):
    # print("combine_counter -> ")
    vehicles, rows_to_skip = get_daily_vehicles_for_counter(data, first_counter_row_num, is_minutely)

    counter_value = get_counter_value(data, first_counter_row_num)
    counter_name = get_counter_name_from_value(counter_value)
    counter_lane = get_counter_lane_from_value(counter_value)

    counter_ix_for_next_counter = first_counter_row_num + rows_to_skip

    try:
        next_counter_value = get_counter_value(data, counter_ix_for_next_counter)
    except Exception as e:
        file_writer.write_combined_values_to_file(vehicles, counter_value, day, month, year, is_minutely)
        return counter_ix_for_next_counter

    # add vehicle numbers together until there is a counter that is not on the same location and direction
    while counter_name == get_counter_name_from_value(next_counter_value) \
            and counter_lane != get_counter_lane_from_value(next_counter_value):

        vehicles_next_counter, rows_to_skip = get_daily_vehicles_for_counter(data, counter_ix_for_next_counter,
                                                                             is_minutely)
        vehicles = combine_vehicle_numbers(vehicles, vehicles_next_counter)

        counter_ix_for_next_counter += rows_to_skip
        try:
            next_counter_value = get_counter_value(data, counter_ix_for_next_counter)
        except Exception as e:
            file_writer.write_combined_values_to_file(vehicles, counter_value, day, month, year, is_minutely)
            return counter_ix_for_next_counter

    file_writer.write_combined_values_to_file(vehicles, counter_value, day, month, year, is_minutely)

    # print("counter_ix_for_next_counter: " + counter_ix_for_next_counter)
    return counter_ix_for_next_counter


def get_day_from_row(data, index):
    try:
        return data.iloc[index, 2]
    except:
        # end of the file
        return -1


def get_month_from_row(data, index):
    try:
        return data.iloc[index, 1]
    except:
        # end of the file
        return -1


def iterate_and_combine_counters_for_one_day(data, day, month, year, counter_start_ix, is_minutely, file_writer):
    # print("iterate_and_combine_counters_for_one_day -> ")

    # loop through all the counters for this day
    previous_day = day
    counter_ix = counter_start_ix

    while day > 0 and day == previous_day:
        # jump to next counter that is on the different location OR direction
        counter_ix = combine_counter(data, counter_ix, day, month, year, is_minutely, file_writer)
        # print("counter_ix: " + str(counter_ix))
        previous_day = day
        day = get_day_from_row(data, counter_ix)

    month = get_month_from_row(data, counter_ix)
    # return -1 if we've reached the end of the year
    return counter_ix, day, month


# combines values of the counters that have the same location and direction,
# count vehicles for different lanes
def combine_lanes(data, is_minutely, year, file_writer):
    counter_start_ix = 0
    day = get_day_from_row(data, counter_start_ix)
    month = get_month_from_row(data, counter_start_ix)
    while day > -1:
        start_time = dt.now()
        d = day
        m = month
        counter_start_ix, day, month = \
            iterate_and_combine_counters_for_one_day(data, day, month, year, counter_start_ix, is_minutely, file_writer)
        print("finished combining lanes for date: " + str(d) + ". " + str(m) + ", time: " + str(dt.now() - start_time))

    file_writer.wrap_it_up(is_minutely)

def get_data_path(year, is_minutely):
    if is_minutely:
        return data_paths[str(year)]["minutely"]
    else:
        return data_paths[str(year)]["hourly"]


def combine_lanes_from_year(year, is_minutely):
    file_writer = FileWriter(year, is_minutely)
    if not test_file:
        try:
            data = pd.read_csv(get_data_path(year, is_minutely), delimiter='\t')
        except Exception as e:
            raise e
    else:
        data = pd.read_csv("./data/random.tab", delimiter='\t', header=None)
    combine_lanes(data, is_minutely, year, file_writer)
    file_writer.save_stuff()

    print("---------------------------------")
    print(file_writer.counter_cols)
    print("---------------------------------")

    # cc = {'0939-1': 2, '0939-2': 3, '0439-1': 4, '0439-2': 5, '0870-1': 6, '0870-2': 7, '0873-1': 8, '0002-1': 9, '0002-2': 10, '0003-1': 11, '0003-2': 12, '0006-1': 13, '0006-2': 14, '0009-1': 15, '0009-2': 16, '0010-1': 17, '0010-2': 18, '0011-1': 19, '0013-1': 20, '0013-2': 21, '0015-1': 22, '0015-2': 23, '0016-1': 24, '0016-2': 25, '0017-1': 26, '0017-2': 27, '0018-1': 28, '0018-2': 29, '0019-1': 30, '0019-2': 31, '0020-1': 32, '0020-2': 33, '0025-1': 34, '0025-2': 35, '0026-1': 36, '0026-2': 37, '0031-1': 38, '0031-2': 39, '0034-1': 40, '0034-2': 41, '0035-1': 42, '0035-2': 43, '0037-1': 44, '0037-2': 45, '0038-1': 46, '0038-2': 47, '0040-1': 48, '0040-2': 49, '0041-1': 50, '0041-2': 51, '0042-1': 52, '0042-2': 53, '0044-1': 54, '0044-2': 55, '0048-1': 56, '0048-2': 57, '0049-1': 58, '0049-2': 59, '0051-1': 60, '0051-2': 61, '0052-1': 62, '0052-2': 63, '0054-1': 64, '0054-2': 65, '0055-1': 66, '0055-2': 67, '0056-1': 68, '0056-2': 69, '0060-1': 70, '0060-2': 71, '0061-1': 72, '0061-2': 73, '0062-1': 74, '0062-2': 75, '0064-1': 76, '0064-2': 77, '0067-1': 78, '0067-2': 79, '0068-1': 80, '0068-2': 81, '0069-1': 82, '0069-2': 83, '0070-1': 84, '0070-2': 85, '0071-1': 86, '0071-2': 87, '0072-1': 88, '0072-2': 89, '0073-1': 90, '0073-2': 91, '0074-1': 92, '0074-2': 93, '0076-1': 94, '0076-2': 95, '0077-1': 96, '0077-2': 97, '0079-1': 98, '0079-2': 99, '0081-1': 100, '0081-2': 101, '0082-1': 102, '0082-2': 103, '0083-1': 104, '0083-2': 105, '0084-1': 106, '0084-2': 107, '0085-1': 108, '0085-2': 109, '0086-1': 110, '0086-2': 111, '0087-1': 112, '0087-2': 113, '0088-1': 114, '0088-2': 115, '0090-1': 116, '0090-2': 117, '0091-1': 118, '0091-2': 119, '0092-1': 120, '0092-2': 121, '0095-1': 122, '0095-2': 123, '0097-1': 124, '0097-2': 125, '0099-1': 126, '0099-2': 127, '0100-1': 128, '0100-2': 129, '0101-1': 130, '0101-2': 131, '0102-1': 132, '0102-2': 133, '0103-1': 134, '0103-2': 135, '0104-1': 136, '0104-2': 137, '0105-1': 138, '0105-2': 139, '0106-1': 140, '0106-2': 141, '0107-1': 142, '0107-2': 143, '0108-1': 144, '0108-2': 145, '0109-1': 146, '0109-2': 147, '0110-1': 148, '0110-2': 149, '0111-1': 150, '0111-2': 151, '0112-1': 152, '0112-2': 153, '0114-1': 154, '0114-2': 155, '0115-1': 156, '0115-2': 157, '0116-1': 158, '0116-2': 159, '0117-1': 160, '0117-2': 161, '0118-1': 162, '0118-2': 163, '0119-1': 164, '0119-2': 165, '0120-1': 166, '0120-2': 167, '0121-1': 168, '0121-2': 169, '0122-1': 170, '0122-2': 171, '0123-1': 172, '0123-2': 173, '0124-1': 174, '0124-2': 175, '0125-1': 176, '0125-2': 177, '0126-1': 178, '0126-2': 179, '0127-1': 180, '0127-2': 181, '0128-1': 182, '0128-2': 183, '0129-1': 184, '0129-2': 185, '0131-1': 186, '0131-2': 187, '0132-1': 188, '0132-2': 189, '0133-1': 190, '0133-2': 191, '0134-1': 192, '0134-2': 193, '0136-1': 194, '0136-2': 195, '0137-1': 196, '0137-2': 197, '0138-1': 198, '0138-2': 199, '0139-1': 200, '0139-2': 201, '0141-1': 202, '0141-2': 203, '0142-1': 204, '0142-2': 205, '0143-1': 206, '0143-2': 207, '0144-1': 208, '0144-2': 209, '0146-1': 210, '0146-2': 211, '0148-1': 212, '0148-2': 213, '0149-1': 214, '0149-2': 215, '0150-1': 216, '0150-2': 217, '0153-1': 218, '0153-2': 219, '0154-1': 220, '0154-2': 221, '0155-1': 222, '0155-2': 223, '0160-1': 224, '0160-2': 225, '0161-1': 226, '0161-2': 227, '0163-1': 228, '0163-2': 229, '0165-1': 230, '0165-2': 231, '0166-1': 232, '0166-2': 233, '0167-1': 234, '0167-2': 235, '0168-1': 236, '0168-2': 237, '0172-1': 238, '0172-2': 239, '0174-2': 240, '0176-1': 241, '0176-2': 242, '0177-1': 243, '0177-2': 244, '0178-1': 245, '0178-2': 246, '0179-2': 247, '0180-1': 248, '0180-2': 249, '0188-1': 250, '0188-2': 251, '0189-1': 252, '0189-2': 253, '0190-1': 254, '0190-2': 255, '0192-1': 256, '0192-2': 257, '0193-1': 258, '0193-2': 259, '0194-1': 260, '0194-2': 261, '0196-1': 262, '0196-2': 263, '0199-1': 264, '0200-1': 265, '0200-2': 266, '0202-1': 267, '0202-2': 268, '0203-1': 269, '0203-2': 270, '0204-1': 271, '0204-2': 272, '0205-1': 273, '0205-2': 274, '0206-1': 275, '0206-2': 276, '0209-1': 277, '0209-2': 278, '0210-1': 279, '0210-2': 280, '0218-1': 281, '0218-2': 282, '0220-1': 283, '0220-2': 284, '0221-1': 285, '0221-2': 286, '0222-1': 287, '0222-2': 288, '0227-1': 289, '0227-2': 290, '0228-2': 291, '0229-1': 292, '0229-2': 293, '0230-1': 294, '0230-2': 295, '0231-1': 296, '0231-2': 297, '0232-1': 298, '0232-2': 299, '0233-1': 300, '0233-2': 301, '0234-1': 302, '0234-2': 303, '0235-1': 304, '0235-2': 305, '0237-1': 306, '0237-2': 307, '0244-1': 308, '0244-2': 309, '0245-1': 310, '0245-2': 311, '0246-1': 312, '0246-2': 313, '0247-1': 314, '0247-2': 315, '0248-1': 316, '0248-2': 317, '0249-1': 318, '0249-2': 319, '0252-1': 320, '0252-2': 321, '0256-1': 322, '0256-2': 323, '0258-1': 324, '0258-2': 325, '0259-1': 326, '0259-2': 327, '0260-1': 328, '0260-2': 329, '0262-1': 330, '0262-2': 331, '0264-1': 332, '0264-2': 333, '0265-1': 334, '0265-2': 335, '0267-1': 336, '0267-2': 337, '0268-1': 338, '0268-2': 339, '0270-1': 340, '0270-2': 341, '0273-1': 342, '0273-2': 343, '0275-1': 344, '0275-2': 345, '0277-1': 346, '0277-2': 347, '0278-1': 348, '0278-2': 349, '0279-1': 350, '0279-2': 351, '0280-1': 352, '0280-2': 353, '0281-1': 354, '0281-2': 355, '0288-1': 356, '0288-2': 357, '0289-1': 358, '0289-2': 359, '0294-1': 360, '0294-2': 361, '0295-1': 362, '0295-2': 363, '0296-1': 364, '0296-2': 365, '0297-1': 366, '0297-2': 367, '0298-1': 368, '0298-2': 369, '0302-1': 370, '0302-2': 371, '0303-1': 372, '0303-2': 373, '0304-1': 374, '0304-2': 375, '0306-1': 376, '0306-2': 377, '0307-1': 378, '0307-2': 379, '0308-1': 380, '0308-2': 381, '0310-1': 382, '0310-2': 383, '0313-1': 384, '0313-2': 385, '0314-1': 386, '0314-2': 387, '0315-1': 388, '0315-2': 389, '0316-1': 390, '0316-2': 391, '0317-1': 392, '0317-2': 393, '0318-1': 394, '0318-2': 395, '0319-1': 396, '0319-2': 397, '0320-1': 398, '0320-2': 399, '0321-1': 400, '0321-2': 401, '0323-1': 402, '0323-2': 403, '0325-1': 404, '0325-2': 405, '0326-1': 406, '0326-2': 407, '0329-1': 408, '0329-2': 409, '0330-1': 410, '0330-2': 411, '0331-1': 412, '0331-2': 413, '0332-1': 414, '0332-2': 415, '0334-1': 416, '0334-2': 417, '0336-1': 418, '0336-2': 419, '0337-1': 420, '0337-2': 421, '0338-1': 422, '0338-2': 423, '0339-1': 424, '0339-2': 425, '0340-1': 426, '0340-2': 427, '0341-1': 428, '0341-2': 429, '0342-1': 430, '0342-2': 431, '0344-1': 432, '0344-2': 433, '0345-1': 434, '0345-2': 435, '0347-1': 436, '0347-2': 437, '0348-1': 438, '0348-2': 439, '0350-1': 440, '0350-2': 441, '0352-1': 442, '0352-2': 443, '0354-1': 444, '0354-2': 445, '0355-1': 446, '0355-2': 447, '0358-1': 448, '0358-2': 449, '0359-1': 450, '0359-2': 451, '0360-1': 452, '0360-2': 453, '0361-1': 454, '0361-2': 455, '0362-1': 456, '0362-2': 457, '0364-1': 458, '0364-2': 459, '0366-1': 460, '0366-2': 461, '0367-1': 462, '0367-2': 463, '0368-1': 464, '0368-2': 465, '0369-1': 466, '0369-2': 467, '0370-1': 468, '0370-2': 469, '0371-1': 470, '0371-2': 471, '0376-1': 472, '0376-2': 473, '0377-1': 474, '0377-2': 475, '0379-1': 476, '0379-2': 477, '0380-1': 478, '0380-2': 479, '0381-1': 480, '0381-2': 481, '0382-1': 482, '0382-2': 483, '0383-1': 484, '0383-2': 485, '0386-1': 486, '0386-2': 487, '0387-1': 488, '0387-2': 489, '0388-1': 490, '0388-2': 491, '0391-1': 492, '0391-2': 493, '0392-1': 494, '0392-2': 495, '0393-1': 496, '0393-2': 497, '0394-1': 498, '0394-2': 499, '0395-1': 500, '0395-2': 501, '0396-1': 502, '0396-2': 503, '0397-1': 504, '0397-2': 505, '0398-1': 506, '0398-2': 507, '0399-1': 508, '0399-2': 509, '0400-1': 510, '0400-2': 511, '0402-1': 512, '0402-2': 513, '0403-1': 514, '0403-2': 515, '0405-1': 516, '0405-2': 517, '0406-1': 518, '0406-2': 519, '0407-1': 520, '0407-2': 521, '0408-1': 522, '0408-2': 523, '0410-1': 524, '0410-2': 525, '0411-1': 526, '0411-2': 527, '0412-1': 528, '0412-2': 529, '0413-1': 530, '0413-2': 531, '0416-1': 532, '0416-2': 533, '0417-1': 534, '0417-2': 535, '0420-1': 536, '0420-2': 537, '0421-1': 538, '0421-2': 539, '0424-1': 540, '0424-2': 541, '0425-1': 542, '0425-2': 543, '0426-1': 544, '0426-2': 545, '0427-1': 546, '0427-2': 547, '0429-1': 548, '0429-2': 549, '0432-1': 550, '0432-2': 551, '0433-1': 552, '0433-2': 553, '0438-1': 554, '0438-2': 555, '0443-1': 556, '0443-2': 557, '0446-1': 558, '0446-2': 559, '0447-1': 560, '0447-2': 561, '0448-1': 562, '0448-2': 563, '0452-1': 564, '0452-2': 565, '0460-1': 566, '0460-2': 567, '0464-1': 568, '0464-2': 569, '0465-1': 570, '0465-2': 571, '0470-1': 572, '0470-2': 573, '0471-1': 574, '0471-2': 575, '0472-1': 576, '0472-2': 577, '0473-1': 578, '0473-2': 579, '0474-1': 580, '0474-2': 581, '0476-1': 582, '0476-2': 583, '0479-1': 584, '0479-2': 585, '0480-1': 586, '0480-2': 587, '0488-1': 588, '0488-2': 589, '0494-1': 590, '0494-2': 591, '0499-1': 592, '0499-2': 593, '0500-1': 594, '0500-2': 595, '0502-1': 596, '0502-2': 597, '0503-1': 598, '0503-2': 599, '0504-1': 600, '0504-2': 601, '0511-1': 602, '0511-2': 603, '0512-1': 604, '0512-2': 605, '0522-1': 606, '0522-2': 607, '0526-1': 608, '0526-2': 609, '0530-1': 610, '0530-2': 611, '0531-1': 612, '0531-2': 613, '0535-1': 614, '0535-2': 615, '0549-1': 616, '0549-2': 617, '0553-1': 618, '0553-2': 619, '0555-1': 620, '0555-2': 621, '0557-1': 622, '0557-2': 623, '0558-1': 624, '0558-2': 625, '0561-1': 626, '0561-2': 627, '0562-1': 628, '0562-2': 629, '0565-1': 630, '0565-2': 631, '0566-1': 632, '0566-2': 633, '0567-1': 634, '0567-2': 635, '0568-1': 636, '0568-2': 637, '0569-1': 638, '0569-2': 639, '0572-1': 640, '0572-2': 641, '0573-1': 642, '0573-2': 643, '0574-1': 644, '0574-2': 645, '0575-1': 646, '0575-2': 647, '0576-1': 648, '0576-2': 649, '0578-1': 650, '0578-2': 651, '0580-1': 652, '0580-2': 653, '0581-1': 654, '0581-2': 655, '0582-1': 656, '0582-2': 657, '0583-1': 658, '0583-2': 659, '0584-1': 660, '0584-2': 661, '0586-1': 662, '0586-2': 663, '0587-1': 664, '0587-2': 665, '0598-1': 666, '0598-2': 667, '0599-1': 668, '0599-2': 669, '0600-1': 670, '0600-2': 671, '0601-1': 672, '0601-2': 673, '0602-2': 674, '0603-1': 675, '0603-2': 676, '0604-1': 677, '0604-2': 678, '0605-1': 679, '0605-2': 680, '0606-1': 681, '0606-2': 682, '0607-1': 683, '0607-2': 684, '0608-1': 685, '0608-2': 686, '0609-1': 687, '0609-2': 688, '0610-1': 689, '0610-2': 690, '0611-1': 691, '0611-2': 692, '0612-1': 693, '0612-2': 694, '0613-1': 695, '0613-2': 696, '0614-1': 697, '0614-2': 698, '0615-1': 699, '0615-2': 700, '0616-1': 701, '0616-2': 702, '0617-1': 703, '0617-2': 704, '0618-1': 705, '0618-2': 706, '0619-1': 707, '0619-2': 708, '0620-1': 709, '0620-2': 710, '0621-1': 711, '0621-2': 712, '0623-1': 713, '0623-2': 714, '0624-1': 715, '0624-2': 716, '0625-1': 717, '0625-2': 718, '0627-1': 719, '0627-2': 720, '0630-1': 721, '0630-2': 722, '0631-1': 723, '0631-2': 724, '0632-1': 725, '0632-2': 726, '0635-1': 727, '0635-2': 728, '0636-1': 729, '0636-2': 730, '0637-1': 731, '0637-2': 732, '0638-1': 733, '0638-2': 734, '0639-1': 735, '0639-2': 736, '0641-1': 737, '0641-2': 738, '0643-1': 739, '0643-2': 740, '0644-1': 741, '0644-2': 742, '0645-1': 743, '0645-2': 744, '0646-1': 745, '0646-2': 746, '0647-1': 747, '0647-2': 748, '0648-1': 749, '0648-2': 750, '0652-1': 751, '0652-2': 752, '0653-1': 753, '0653-2': 754, '0654-1': 755, '0654-2': 756, '0662-1': 757, '0662-2': 758, '0663-1': 759, '0663-2': 760, '0664-1': 761, '0664-2': 762, '0665-1': 763, '0665-2': 764, '0667-1': 765, '0667-2': 766, '0669-1': 767, '0669-2': 768, '0670-1': 769, '0670-2': 770, '0671-1': 771, '0671-2': 772, '0673-1': 773, '0673-2': 774, '0674-1': 775, '0674-2': 776, '0676-1': 777, '0676-2': 778, '0678-1': 779, '0678-2': 780, '0679-1': 781, '0679-2': 782, '0681-1': 783, '0681-2': 784, '0683-1': 785, '0683-2': 786, '0684-1': 787, '0684-2': 788, '0686-1': 789, '0686-2': 790, '0687-1': 791, '0687-2': 792, '0688-1': 793, '0688-2': 794, '0689-1': 795, '0689-2': 796, '0690-1': 797, '0690-2': 798, '0691-1': 799, '0691-2': 800, '0692-1': 801, '0692-2': 802, '0693-1': 803, '0693-2': 804, '0694-1': 805, '0694-2': 806, '0695-1': 807, '0696-1': 808, '0696-2': 809, '0697-1': 810, '0697-2': 811, '0699-1': 812, '0699-2': 813, '0706-1': 814, '0706-2': 815, '0707-1': 816, '0707-2': 817, '0708-1': 818, '0710-1': 819, '0710-2': 820, '0711-1': 821, '0711-2': 822, '0715-1': 823, '0715-2': 824, '0716-1': 825, '0716-2': 826, '0717-1': 827, '0717-2': 828, '0718-1': 829, '0718-2': 830, '0725-1': 831, '0725-2': 832, '0735-1': 833, '0735-2': 834, '0736-1': 835, '0736-2': 836, '0738-1': 837, '0738-2': 838, '0739-1': 839, '0739-2': 840, '0740-1': 841, '0740-2': 842, '0741-1': 843, '0741-2': 844, '0742-1': 845, '0742-2': 846, '0743-1': 847, '0743-2': 848, '0744-1': 849, '0744-2': 850, '0746-1': 851, '0746-2': 852, '0747-1': 853, '0747-2': 854, '0748-1': 855, '0748-2': 856, '0751-1': 857, '0751-2': 858, '0752-1': 859, '0752-2': 860, '0754-1': 861, '0754-2': 862, '0755-1': 863, '0755-2': 864, '0756-1': 865, '0756-2': 866, '0757-1': 867, '0757-2': 868, '0776-1': 869, '0776-2': 870, '0777-1': 871, '0777-2': 872, '0781-1': 873, '0781-2': 874, '0782-1': 875, '0782-2': 876, '0792-1': 877, '0792-2': 878, '0793-1': 879, '0793-2': 880, '0794-1': 881, '0794-2': 882, '0799-1': 883, '0799-2': 884, '0802-1': 885, '0802-2': 886, '0803-1': 887, '0803-2': 888, '0806-1': 889, '0806-2': 890, '0808-1': 891, '0808-2': 892, '0810-1': 893, '0810-2': 894, '0812-1': 895, '0812-2': 896, '0813-1': 897, '0813-2': 898, '0815-1': 899, '0815-2': 900, '0821-1': 901, '0821-2': 902, '0822-1': 903, '0822-2': 904, '0824-1': 905, '0824-2': 906, '0826-1': 907, '0826-2': 908, '0828-1': 909, '0828-2': 910, '0830-1': 911, '0830-2': 912, '0832-1': 913, '0832-2': 914, '0834-1': 915, '0834-2': 916, '0835-1': 917, '0835-2': 918, '0836-1': 919, '0836-2': 920, '0837-1': 921, '0837-2': 922, '0838-1': 923, '0838-2': 924, '0839-1': 925, '0839-2': 926, '0840-1': 927, '0840-2': 928, '0841-1': 929, '0841-2': 930, '0846-1': 931, '0846-2': 932, '0848-1': 933, '0848-2': 934, '0850-1': 935, '0850-2': 936, '0853-1': 937, '0853-2': 938, '0854-1': 939, '0854-2': 940, '0855-1': 941, '0855-2': 942, '0856-1': 943, '0856-2': 944, '0860-1': 945, '0860-2': 946, '0861-1': 947, '0861-2': 948, '0865-1': 949, '0865-2': 950, '0866-1': 951, '0866-2': 952, '0873-2': 953, '0875-1': 954, '0875-2': 955, '0877-1': 956, '0877-2': 957, '0878-1': 958, '0878-2': 959, '0880-1': 960, '0880-2': 961, '0881-1': 962, '0881-2': 963, '0884-1': 964, '0884-2': 965, '0885-1': 966, '0885-2': 967, '0887-1': 968, '0887-2': 969, '0889-1': 970, '0889-2': 971, '0890-1': 972, '0890-2': 973, '0892-1': 974, '0893-1': 975, '0893-2': 976, '0894-1': 977, '0895-2': 978, '0896-1': 979, '0896-2': 980, '0897-1': 981, '0897-2': 982, '0900-1': 983, '0900-2': 984, '0901-1': 985, '0901-2': 986, '0902-1': 987, '0902-2': 988, '0903-1': 989, '0903-2': 990, '0904-1': 991, '0904-2': 992, '0905-1': 993, '0905-2': 994, '0906-1': 995, '0906-2': 996, '0907-1': 997, '0907-2': 998, '0908-1': 999, '0908-2': 1000, '0909-1': 1001, '0909-2': 1002, '0910-1': 1003, '0910-2': 1004, '0911-1': 1005, '0911-2': 1006, '0912-1': 1007, '0912-2': 1008, '0913-1': 1009, '0913-2': 1010, '0914-1': 1011, '0914-2': 1012, '0915-1': 1013, '0915-2': 1014, '0916-1': 1015, '0916-2': 1016, '0917-1': 1017, '0917-2': 1018, '0918-1': 1019, '0918-2': 1020, '0919-1': 1021, '0919-2': 1022, '0920-1': 1023, '0920-2': 1024, '0921-1': 1025, '0921-2': 1026, '0922-1': 1027, '0922-2': 1028, '0923-1': 1029, '0923-2': 1030, '0924-1': 1031, '0924-2': 1032, '0925-1': 1033, '0925-2': 1034, '0926-1': 1035, '0926-2': 1036, '0927-1': 1037, '0927-2': 1038, '0928-1': 1039, '0928-2': 1040, '0929-1': 1041, '0929-2': 1042, '0930-1': 1043, '0930-2': 1044, '0931-1': 1045, '0931-2': 1046, '0932-1': 1047, '0932-2': 1048, '0934-1': 1049, '0934-2': 1050, '0935-1': 1051, '0935-2': 1052, '0936-1': 1053, '0936-2': 1054, '0937-1': 1055, '0937-2': 1056, '1001-1': 1057, '1002-1': 1058, '1003-1': 1059, '1004-1': 1060, '1005-1': 1061, '1005-2': 1062, '1006-1': 1063, '1006-2': 1064, '1007-1': 1065, '1007-2': 1066, '1008-1': 1067, '1008-2': 1068, '1009-1': 1069, '1009-2': 1070, '1010-1': 1071, '1010-2': 1072, '1011-1': 1073, '1011-2': 1074, '1012-1': 1075, '1013-1': 1076, '1013-2': 1077, '1014-1': 1078, '1014-2': 1079, '1015-1': 1080, '1015-2': 1081, '1016-1': 1082, '1016-2': 1083, '1017-1': 1084, '1017-2': 1085, '1018-1': 1086, '1018-2': 1087, '1019-1': 1088, '1019-2': 1089, '1020-1': 1090, '1020-2': 1091, '1021-1': 1092, '1021-2': 1093, '1022-1': 1094, '1022-2': 1095, '1023-1': 1096, '1023-2': 1097, '1025-1': 1098, '1025-2': 1099, '1026-1': 1100, '1027-1': 1101, '1027-2': 1102, '1028-1': 1103, '1029-1': 1104, '1029-2': 1105, '1030-1': 1106, '1030-2': 1107, '1031-1': 1108, '1031-2': 1109, '1032-1': 1110, '1032-2': 1111, '1033-1': 1112, '1033-2': 1113, '1034-1': 1114, '1034-2': 1115, '1035-1': 1116, '1036-1': 1117, '1036-2': 1118, '0441-1': 1119, '0441-2': 1120, '0156-1': 1121, '0156-2': 1122, '0201-1': 1123, '0201-2': 1124, '0622-1': 1125, '0622-2': 1126, '1039-1': 1127, '1039-2': 1128, '0994-1': 1129, '0994-2': 1130, '1035-2': 1131, '1037-2': 1132, '1044-2': 1133}

    replace_null_values(file_writer.counter_cols, year, is_minutely)

#############################################################################################


def combine_years(is_minutely):
    data_2019 = pd.read_excel("./data/traffic_2019_combined.xlsx", header=0)
    data_2020 = pd.read_excel("./data/traffic_2020_combined.xlsx", header=0)
    data_2021 = pd.read_excel("./data/traffic_2021_combined.xlsx", header=0)

    print("data loaded")

    values_for_day = get_values_for_day(is_minutely)

    workbook = xlsxwriter.Workbook('./data/data_combined.xlsx')
    combined = workbook.add_worksheet()

    header_lines = {}

    first_empty_column_ix = 2

    for i in range(2, len(data_2019.columns)):
        combined.write(0, i, data_2019.columns[i])
        column_values = data_2019.iloc[:, i].values
        column_values = [x if not np.isnan(x) else "" for x in column_values]
        combined.write_column(1, i, column_values)
        header_lines[data_2019.columns[i]] = {"start_ix": first_empty_column_ix, "length": len(column_values)+1}
        first_empty_column_ix += 1

    print("2019 finished")

    # check
    for key, value in header_lines.items():
        if value["length"] != (365 * values_for_day + 1):
            raise Exception("check this")

    for i in range(2, len(data_2020.columns)):

        if data_2020.columns[i] in header_lines:
            column_values = data_2020.iloc[:, i].values
            column_values = [x if not np.isnan(x) else "" for x in column_values]
            combined.write_column(header_lines[data_2020.columns[i]]["length"], header_lines[data_2020.columns[i]]["start_ix"], column_values)
            header_lines[data_2020.columns[i]]["length"] = header_lines[data_2020.columns[i]]["length"] + len(column_values)
        else:
            combined.write_column(first_empty_column_ix, 0,
                                  data_2020.columns[i])
            column_values = data_2020.iloc[:, i].values
            column_values = [x if not np.isnan(x) else "" for x in column_values]
            combined.write_column(365*values_for_day+1, first_empty_column_ix,
                                  column_values)
            header_lines[data_2020.columns[i]] = {"start_ix": first_empty_column_ix, "length": 365*values_for_day+1+len(column_values)}
            first_empty_column_ix += 1

    print("2020 finished")

    for i in range(2, len(data_2021.columns)):

        if data_2021.columns[i] in header_lines:
            column_values = data_2021.iloc[:, i].values
            column_values = [x if not np.isnan(x) else "" for x in column_values]
            combined.write_column(header_lines[data_2021.columns[i]]["length"], header_lines[data_2021.columns[i]]["start_ix"], column_values)
            header_lines[data_2021.columns[i]]["length"] = header_lines[data_2021.columns[i]]["length"] + len(column_values)
        else:
            combined.write_column(first_empty_column_ix, 0,
                                  data_2021.columns[i])
            column_values = data_2021.iloc[:, i].values
            column_values = [x if not np.isnan(x) else "" for x in column_values]
            combined.write_column(365*values_for_day+366*values_for_day+1, first_empty_column_ix,
                                  column_values)

            header_lines[data_2021.columns[i]] = {"start_ix": first_empty_column_ix, "length": 365*values_for_day+366*values_for_day+1+len(column_values)}
            first_empty_column_ix += 1

    print("2021 finished")

    days_2019 = data_2019.iloc[:, 0].values
    days_2019 = [value if not str(value) == 'nan' else '' for value in days_2019]
    hours_2019 = data_2019.iloc[:, 1].values
    hours_2019 = [value if not str(value) == 'nan' else '' for value in hours_2019]

    days_2020 = data_2020.iloc[:, 0].values
    days_2020 = [value if not str(value) == 'nan' else '' for value in days_2020]
    hours_2020 = data_2020.iloc[:, 1].values
    hours_2020 = [value if not str(value) == 'nan' else '' for value in hours_2020]

    days_2021 = data_2021.iloc[:, 0].values
    days_2021 = [value if not str(value) == 'nan' else '' for value in days_2021]
    hours_2021 = data_2021.iloc[:, 1].values
    hours_2021 = [value if not str(value) == 'nan' else '' for value in hours_2021]

    combined.write_column(1, 0, days_2019)
    combined.write_column(1, 1, hours_2019)
    combined.write_column(365*values_for_day+1, 0, days_2020)
    combined.write_column(365*values_for_day+1, 1, hours_2020)
    combined.write_column(365*values_for_day+1+366*values_for_day, 0, days_2021)
    combined.write_column(365*values_for_day+1+366*values_for_day, 1, hours_2021)

    workbook.close()


test_file = False
# combine_lanes_from_year(2019, is_minutely=False)
combine_years(False)








